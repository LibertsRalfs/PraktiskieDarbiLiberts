from openpyxl import load_workbook

 #Atver excela failu
wb = load_workbook('tests/test1.xlsx')
ws = wb.active
total = 0

for row in ws.iter_rows(min_row=2, values_only=True):
        #Parliecnās vai dati ir tiešām skaitļos
     if isinstance(row[1], (int, float)) and isinstance(row[2], (int, float)):
        monthly_salary = row[1] * row[2]
        if monthly_salary > 3000:
            total += 1

         #Aizver excel failu
wb.close()

    #Darbinieku skaits kuru alga ir lielāka par 3000
print("Darbinieku skaits kuru alga ir lielāka par 3000 ir:")
print(total)




